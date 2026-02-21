import gitlab
from openpyxl import load_workbook

# ==========================
# CONFIG
# ==========================
GITLAB_URL = "https://gitlab.com"
PRIVATE_TOKEN = "YOUR_PRIVATE_TOKEN"
EXCEL_FILE = "your_excel_file.xlsx"

ASSIGNEE_VALUE = "chandresh version fix"

# ==========================
# CONNECT GITLAB
# ==========================
gl = gitlab.Gitlab(GITLAB_URL, private_token=PRIVATE_TOKEN)

# ==========================
# LOAD EXCEL
# ==========================
wb = load_workbook(EXCEL_FILE)
ws = wb["Version"]

# Get header indexes dynamically
headers = {}
for col in range(1, ws.max_column + 1):
    headers[ws.cell(row=1, column=col).value] = col

# Required columns
assignee_col = headers["Version assignee"]
repo_col = headers["Repsitory URL"]
mr_main_col = headers["MR from version to main"]
mr_dev_col = headers["MR from version to DEV"]
mr_qa_col = headers["MR from version to QA"]

# ==========================
# LOOP ROWS
# ==========================
for row in range(2, ws.max_row + 1):

    assignee = ws.cell(row=row, column=assignee_col).value

    if assignee != ASSIGNEE_VALUE:
        continue

    repo_url = ws.cell(row=row, column=repo_col).value

    if not repo_url:
        continue

    # Extract project path from repo URL
    # Example: https://gitlab.com/group/project.git
    project_path = repo_url.replace(GITLAB_URL + "/", "").replace(".git", "")
    project = gl.projects.get(project_path)

    # ==========================
    # CREATE MR IF EMPTY
    # ==========================

    def create_mr_if_missing(target_branch, column_index):
        cell = ws.cell(row=row, column=column_index)

        if cell.value:  # already has link
            return

        mr = project.mergerequests.create({
            "source_branch": "version",
            "target_branch": target_branch,
            "title": f"From version to {target_branch}",
            "description": f"From version to {target_branch}",
            "remove_source_branch": False,
            "squash": False,
        })

        # Insert clickable hyperlink
        cell.value = f"MR to {target_branch}"
        cell.hyperlink = mr.web_url
        cell.style = "Hyperlink"

        print(f"Created MR for row {row}: {mr.web_url}")

    create_mr_if_missing("main", mr_main_col)
    create_mr_if_missing("DEV", mr_dev_col)
    create_mr_if_missing("QA", mr_qa_col)

# ==========================
# SAVE FILE
# ==========================
wb.save(EXCEL_FILE)

print("Process completed successfully.")